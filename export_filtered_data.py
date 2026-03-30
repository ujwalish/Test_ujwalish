import pandas as pd
import openpyxl.utils
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the data
try:
    df = pd.read_csv("27-Mar-2026 Oracle Report!.csv", encoding='latin1')
except Exception:
    try:
        df = pd.read_csv("27-Mar-2026 Oracle Report!.csv", encoding='iso-8859-1')
    except:
        df = pd.read_csv("27-Mar-2026 Oracle Report!.csv", encoding='cp1252')

print("=" * 120)
print("EXPORTING FILTERED DATA TO EXCEL")
print("=" * 120)

# Apply filters
filtered_df = df[
    (df['Account Status'] == 'Active') &
    (df['Region Of Branch'].isin(['Region-3 Office', 'Region-4 Office']))
].copy()

print(f"\n✓ Applied Filters:")
print(f"  - Account Status: Active")
print(f"  - Region Of Branch: Region-3 Office OR Region-4 Office")
print(f"\n✓ Total Records after filtering: {len(filtered_df)}")
print(f"✓ Unique Sectors: {filtered_df['Sector'].nunique()}")
print(f"✓ Unique Organizations: {filtered_df['Subscription Name'].nunique()}")

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=11)

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Sheet 1: Summary
ws_summary = wb.create_sheet("Summary", 0)
ws_summary.column_dimensions['A'].width = 40
ws_summary.column_dimensions['B'].width = 25

# Title
ws_summary['A1'] = "ACTIVE ACCOUNTS - REGION 3 & 4 SUMMARY"
ws_summary['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws_summary.merge_cells('A1:B1')

ws_summary['A2'] = "Filter: Account Status = Active, Region = Region-3 Office or Region-4 Office"
ws_summary['A2'].font = Font(italic=True, size=10)
ws_summary.merge_cells('A2:B2')

ws_summary['A3'] = "Generated: 29-Mar-2026"
ws_summary['A3'].font = Font(size=9, color="808080")

# Summary metrics
row = 5
ws_summary[f'A{row}'] = "SUMMARY STATISTICS"
ws_summary[f'A{row}'].font = subheader_font
ws_summary[f'A{row}'].fill = subheader_fill
ws_summary.merge_cells(f'A{row}:B{row}')

row += 2
metrics = [
    ("Total Records", len(filtered_df)),
    ("Unique Sectors", filtered_df['Sector'].nunique()),
    ("Unique Organizations", filtered_df['Subscription Name'].nunique()),
    ("Unique Circuits (User Subscription IDs)", filtered_df['User Subscription Id'].nunique()),
]

for metric_name, metric_value in metrics:
    ws_summary[f'A{row}'] = metric_name
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = metric_value
    ws_summary[f'B{row}'].font = Font(size=11, bold=True, color="1F4E78")
    row += 1

# Regional breakdown
row += 2
ws_summary[f'A{row}'] = "REGIONAL BREAKDOWN"
ws_summary[f'A{row}'].font = subheader_font
ws_summary[f'A{row}'].fill = subheader_fill
ws_summary.merge_cells(f'A{row}:B{row}')

row += 1
for region in ['Region-3 Office', 'Region-4 Office']:
    region_count = len(filtered_df[filtered_df['Region Of Branch'] == region])
    region_orgs = filtered_df[filtered_df['Region Of Branch'] == region]['Subscription Name'].nunique()
    
    ws_summary[f'A{row}'] = region
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'] = f"{region_count} records | {region_orgs} organizations"
    row += 1

# Top sectors
row += 2
ws_summary[f'A{row}'] = "TOP SECTORS (by Circuit Count)"
ws_summary[f'A{row}'].font = subheader_font
ws_summary[f'A{row}'].fill = subheader_fill
ws_summary.merge_cells(f'A{row}:B{row}')

row += 1
top_sectors = filtered_df.groupby('Sector').agg({
    'User Subscription Id': 'count',
    'Subscription Name': 'nunique'
}).sort_values('User Subscription Id', ascending=False).head(10)

for sector, data in top_sectors.iterrows():
    ws_summary[f'A{row}'] = sector
    ws_summary[f'B{row}'] = f"{int(data['User Subscription Id'])} circuits | {int(data['Subscription Name'])} orgs"
    row += 1

# Sheet 2: Detailed Data
ws_data = wb.create_sheet("Detailed Data", 1)

# Get all columns except Amount and Dealt Amount
exclude_cols = ['Amount', 'Dealt Amount']
available_cols = [col for col in df.columns if col not in exclude_cols]

# Set uniform column width for all columns
for col_idx in range(1, len(available_cols) + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws_data.column_dimensions[col_letter].width = 18

filtered_display = filtered_df[available_cols].copy()

# Write header
for c_idx, col_name in enumerate(available_cols, 1):
    cell = ws_data.cell(row=1, column=c_idx)
    cell.value = col_name
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Write data
for r_idx, row_data in enumerate(dataframe_to_rows(filtered_display, index=False, header=False), 2):
    for c_idx, value in enumerate(row_data, 1):
        cell = ws_data.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal="left", wrap_text=False)

# Freeze header row
ws_data.freeze_panes = "A2"

# Sheet 3: Sector Wise Summary
ws_sector = wb.create_sheet("Sector Summary", 2)
ws_sector.column_dimensions['A'].width = 30
ws_sector.column_dimensions['B'].width = 20
ws_sector.column_dimensions['C'].width = 20

# Header
ws_sector['A1'] = "SECTOR WISE SUMMARY"
ws_sector['A1'].font = subheader_font
ws_sector['A1'].fill = subheader_fill
ws_sector.merge_cells('A1:C1')

# Column headers
headers = ['Sector', 'Circuit Count', 'Organization Count']
for c_idx, header in enumerate(headers, 1):
    cell = ws_sector.cell(row=2, column=c_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Sector summary data
sector_summary = filtered_df.groupby('Sector').agg({
    'User Subscription Id': 'count',
    'Subscription Name': 'nunique'
}).sort_values('User Subscription Id', ascending=False).reset_index()

for r_idx, (_, row_data) in enumerate(sector_summary.iterrows(), 3):
    ws_sector.cell(row=r_idx, column=1).value = row_data['Sector']
    ws_sector.cell(row=r_idx, column=2).value = row_data['User Subscription Id']
    ws_sector.cell(row=r_idx, column=3).value = row_data['Subscription Name']
    
    for col in range(1, 4):
        ws_sector.cell(row=r_idx, column=col).border = border
        if col > 1:
            ws_sector.cell(row=r_idx, column=col).alignment = Alignment(horizontal="right")

# Save workbook
output_file = "Active_Region3_Region4_Export.xlsx"
wb.save(output_file)

print(f"\n✓ Excel file saved to: {output_file}")
print(f"  Location: c:\\dashboard_project\\{output_file}")
print(f"\nWorkbook Sheets:")
print(f"  1. Summary - Key metrics and regional breakdown")
print(f"  2. Detailed Data - Complete list of filtered records")
print(f"  3. Sector Summary - Sector-wise aggregation")
print("\n" + "=" * 120)
